"""
招生计划数据比对与转换工具 - 核心功能模块
功能：
1. 比对1：招生计划 vs 专业分
2. 比对2：招生计划 vs 院校分
3. 未匹配数据转换为专业分格式导出
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from io import BytesIO
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ==================== 数据加载函数 ====================

def load_excel_file(file_path):
    """加载Excel文件并返回DataFrame"""
    try:
        df = pd.read_excel(file_path, sheet_name=0)
        logger.info(f"成功加载文件，共 {len(df)} 条记录，{len(df.columns)} 列")
        return df
    except Exception as e:
        logger.error(f"加载文件失败: {e}")
        raise


def load_excel_from_bytes(file_bytes):
    """从字节流加载Excel文件"""
    try:
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=0)
        logger.info(f"成功加载文件，共 {len(df)} 条记录，{len(df.columns)} 列")
        return df
    except Exception as e:
        logger.error(f"加载文件失败: {e}")
        raise


# ==================== 关键字生成函数 ====================

def generate_plan_score_key(row):
    """
    生成招生计划 vs 专业分的比对关键字
    关键字字段：年份、省份、学校、科类、批次、专业、层次、专业组代码
    """
    try:
        key_parts = [
            str(row.get('年份', '')).strip(),
            str(row.get('省份', '')).strip(),
            str(row.get('学校', '')).strip(),
            str(row.get('科类', '')).strip(),
            str(row.get('批次', '')).strip(),
            str(row.get('专业', '')).strip(),
            str(row.get('层次', '')).strip(),
            str(row.get('专业组代码', '')).strip()
        ]
        return '|'.join(key_parts)
    except Exception as e:
        logger.error(f"生成关键字失败: {e}")
        return '|'.join([''] * 8)


def generate_plan_college_key(row):
    """
    生成招生计划 vs 院校分的比对关键字
    关键字字段：年份、省份、学校、科类、批次、专业组代码
    """
    try:
        key_parts = [
            str(row.get('年份', '')).strip(),
            str(row.get('省份', '')).strip(),
            str(row.get('学校', '')).strip(),
            str(row.get('科类', '')).strip(),
            str(row.get('批次', '')).strip(),
            str(row.get('专业组代码', '')).strip()
        ]
        return '|'.join(key_parts)
    except Exception as e:
        logger.error(f"生成关键字失败: {e}")
        return '|'.join([''] * 6)


# ==================== 数据比对函数 ====================

def compare_plan_vs_score(plan_df, score_df):
    """
    比对1：招生计划 vs 专业分
    返回：比对结果列表
    """
    results = []
    
    # 为专业分建立关键字索引
    score_keys = set()
    for idx, row in score_df.iterrows():
        key = generate_plan_score_key(row)
        score_keys.add(key)
    
    logger.info(f"专业分关键字数: {len(score_keys)}")
    
    # 比对招生计划
    for idx, row in plan_df.iterrows():
        key = generate_plan_score_key(row)
        exists = key in score_keys
        
        result = {
            'index': idx + 1,
            'original_index': idx,
            'key_fields': {
                '年份': row.get('年份', ''),
                '省份': row.get('省份', ''),
                '学校': row.get('学校', ''),
                '科类': row.get('科类', ''),
                '批次': row.get('批次', ''),
                '专业': row.get('专业', ''),
                '层次': row.get('层次', ''),
                '专业组代码': row.get('专业组代码', '')
            },
            'exists': exists,
            'other_info': {
                '招生人数': row.get('招生人数', ''),
                '学费': row.get('学费', ''),
                '学制': row.get('学制', ''),
                '专业代码': row.get('专业代码', ''),
                '招生代码': row.get('招生代码', ''),
                '数据来源': row.get('数据来源', ''),
                '备注': row.get('备注', ''),
                '招生类型': row.get('招生类型', ''),
                '专业组选科要求': row.get('专业组选科要求', ''),
                '专业选科要求': row.get('专业选科要求(新高考专业省份)', '')
            },
            'raw_data': row
        }
        results.append(result)
    
    logger.info(f"比对1完成: 总记录 {len(plan_df)}, 匹配 {sum(1 for r in results if r['exists'])}, "
                f"未匹配 {sum(1 for r in results if not r['exists'])}")
    
    return results


def compare_plan_vs_college(plan_df, college_df):
    """
    比对2：招生计划 vs 院校分
    返回：比对结果列表
    """
    results = []
    
    # 为院校分建立关键字索引
    college_keys = set()
    for idx, row in college_df.iterrows():
        key = generate_plan_college_key(row)
        college_keys.add(key)
    
    logger.info(f"院校分关键字数: {len(college_keys)}")
    
    # 比对招生计划
    for idx, row in plan_df.iterrows():
        key = generate_plan_college_key(row)
        exists = key in college_keys
        
        result = {
            'index': idx + 1,
            'original_index': idx,
            'key_fields': {
                '年份': row.get('年份', ''),
                '省份': row.get('省份', ''),
                '学校': row.get('学校', ''),
                '科类': row.get('科类', ''),
                '批次': row.get('批次', ''),
                '专业组代码': row.get('专业组代码', '')
            },
            'exists': exists,
            'other_info': {
                '专业': row.get('专业', ''),
                '层次': row.get('层次', ''),
                '招生人数': row.get('招生人数', ''),
                '学费': row.get('学费', ''),
                '学制': row.get('学制', ''),
                '专业代码': row.get('专业代码', ''),
                '招生代码': row.get('招生代码', ''),
                '数据来源': row.get('数据来源', ''),
                '备注': row.get('备注', ''),
                '招生类型': row.get('招生类型', ''),
                '专业组选科要求': row.get('专业组选科要求', ''),
                '专业选科要求': row.get('专业选科要求(新高考专业省份)', '')
            },
            'raw_data': row
        }
        results.append(result)
    
    logger.info(f"比对2完成: 总记录 {len(plan_df)}, 匹配 {sum(1 for r in results if r['exists'])}, "
                f"未匹配 {sum(1 for r in results if not r['exists'])}")
    
    return results


# ==================== 统计函数 ====================

def get_comparison_stats(results):
    """获取比对结果统计信息"""
    total = len(results)
    matched = sum(1 for r in results if r['exists'])
    unmatched = total - matched
    match_rate = (matched / total * 100) if total > 0 else 0
    
    return {
        'total': total,
        'matched': matched,
        'unmatched': unmatched,
        'match_rate': f"{match_rate:.2f}%"
    }


def get_unique_provinces(results):
    """从比对结果中提取唯一的省份列表"""
    provinces = set()
    for result in results:
        province = result['key_fields'].get('省份', '')
        if province:
            provinces.add(str(province).strip())
    return sorted(list(provinces))


def get_unique_batches(results):
    """从比对结果中提取唯一的批次列表"""
    batches = set()
    for result in results:
        batch = result['key_fields'].get('批次', '')
        if batch:
            batches.add(str(batch).strip())
    return sorted(list(batches))


# ==================== 数据转换函数 ====================

def get_first_subject(category):
    """
    获取首选科目：根据招生科类的第一个字
    """
    category_str = str(category).strip()
    if not category_str:
        return ''
    
    # 直接取第一个字作为首选科目
    first_char = category_str[0]
    subject_map = {
        '物': '物',
        '历': '历',
        '文': '文',
        '理': '理',
        '综': '综'
    }
    return subject_map.get(first_char, first_char)


def convert_level(level):
    """转换层次字段"""
    level_str = str(level).strip().lower()
    
    conversion_map = {
        '本科': '本科',
        'undergraduate': '本科',
        '专科': '专科（高职）',
        'vocational': '专科（高职）',
        '高职': '专科（高职）',
        '职高': '专科（高职）'
    }
    
    for key, value in conversion_map.items():
        if key in level_str:
            return value
    
    return level


def extract_required_subjects(text):
    """
    提取必选科目
    处理格式如："物化生（3科必选）"、"物理、化学、生物"等
    """
    if not text:
        return []
    
    text_str = str(text).strip()
    
    # 科目字符
    subjects = ['物', '化', '生', '历', '地', '政', '技']
    found_subjects = []
    
    for subject in subjects:
        if subject in text_str:
            found_subjects.append(subject)
    
    return found_subjects


def convert_selection_requirement(group_requirement, major_requirement=''):
    """
    转换选科要求
    """
    group_req_str = str(group_requirement).strip()
    
    if not group_req_str or group_req_str.lower() == 'nan':
        return '不限科目专业组'
    
    # 检查是否包含"必选"关键词
    if '必选' in group_req_str:
        return '单科、多科均需选考'
    
    # 检查是否包含"不限"
    if '不限' in group_req_str:
        return '不限科目专业组'
    
    # 检查是否包含多门
    if '多门' in group_req_str or '或' in group_req_str:
        return '多门选考'
    
    # 默认为多门选考
    return '多门选考'


def convert_data_to_score_format(unmatched_data, plan_df_original):
    """
    将未匹配的招生计划数据转换为专业分导入模板格式
    """
    converted_data = []
    
    headers = [
        '学校名称', '省份', '招生专业', '专业方向（选填）', '专业备注（选填）',
        '一级层次', '招生科类', '招生批次', '招生类型（选填）', '最高分',
        '最低分', '平均分', '最低分位次（选填）', '招生人数（选填）',
        '数据来源', '专业组代码', '首选科目', '选科要求', '次选科目',
        '专业代码', '招生代码', '最低分数区间低', '最低分数区间高',
        '最低分数区间位次低', '最低分数区间位次高', '录取人数（选填）'
    ]
    
    for item in unmatched_data:
        try:
            original_index = item['original_index']
            raw_row = plan_df_original.iloc[original_index]
            
            # 提取选科要求
            group_req = raw_row.get('专业组选科要求', '')
            major_req = raw_row.get('专业选科要求(新高考专业省份)', '')
            
            # 提取必选科目
            required_subjects = extract_required_subjects(group_req)
            
            # 次选科目：如果有必选科目，取第一个
            second_subject = required_subjects[0] if required_subjects else ''
            
            converted_row = {
                '学校名称': raw_row.get('学校', ''),
                '省份': raw_row.get('省份', ''),
                '招生专业': raw_row.get('专业', ''),
                '专业方向（选填）': '',
                '专业备注（选填）': raw_row.get('备注', ''),
                '一级层次': convert_level(raw_row.get('层次', '')),
                '招生科类': raw_row.get('科类', ''),
                '招生批次': raw_row.get('批次', ''),
                '招生类型（选填）': raw_row.get('招生类型', ''),
                '最高分': '',
                '最低分': '',
                '平均分': '',
                '最低分位次（选填）': '',
                '招生人数（选填）': raw_row.get('招生人数', ''),
                '数据来源': raw_row.get('数据来源', ''),
                '专业组代码': raw_row.get('专业组代码', ''),
                '首选科目': get_first_subject(raw_row.get('科类', '')),
                '选科要求': convert_selection_requirement(group_req, major_req),
                '次选科目': second_subject,
                '专业代码': raw_row.get('专业代码', ''),
                '招生代码': raw_row.get('招生代码', ''),
                '最低分数区间低': '',
                '最低分数区间高': '',
                '最低分数区间位次低': '',
                '最低分数区间位次高': '',
                '录取人数（选填）': ''
            }
            
            converted_data.append(converted_row)
        except Exception as e:
            logger.error(f"转换数据失败 (索引 {item['original_index']}): {e}")
            continue
    
    return converted_data


# ==================== 导出函数 ====================

def export_results_to_excel(results, filename, is_unmatched=False):
    """
    导出比对结果到Excel文件
    """
    try:
        # 过滤数据
        if is_unmatched:
            results = [r for r in results if not r['exists']]
        
        # 创建DataFrame
        data_for_export = []
        for result in results:
            row = {
                '序号': result['index'],
                '匹配状态': '✓ 匹配' if result['exists'] else '✗ 未匹配',
                **result['key_fields'],
                **result['other_info']
            }
            data_for_export.append(row)
        
        df = pd.DataFrame(data_for_export)
        
        # 创建Excel工作簿
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='比对结果', index=False)
            
            # 获取工作表
            workbook = writer.book
            worksheet = writer.sheets['比对结果']
            
            # 设置列宽
            for column in worksheet.columns:
                max_length = 12
                column_letter = column[0].column_letter
                worksheet.column_dimensions[column_letter].width = max_length
            
            # 设置头部样式
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        output.seek(0)
        return output.getvalue()
    
    except Exception as e:
        logger.error(f"导出失败: {e}")
        raise


def export_converted_data_to_excel(converted_data, admission_year=''):
    """
    导出转换后的数据为专业分导入模板格式
    """
    try:
        # Excel工作簿
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = '专业分数据'
        
        # 第1行：说明文本（合并A1~Y1）
        remark_text = (
            "1.省份：必须填写各省份简称，例如：北京、内蒙古，不能带有市、省、自治区、空格、特殊字符等 "
            "2.科类：浙江、上海限定\"综合、艺术类、体育类\"，内蒙古限定\"文科、理科、蒙授文科、蒙授理科、艺术类、艺术文、艺术理、体育类、体育文、体育理、蒙授艺术、蒙授体育\"，其他省份限定\"文科、理科、艺术类、艺术文、艺术理、体育类、体育文、体育理\" "
            "3.批次：河北、内蒙古等省份限定本科提前批、本科一批、本科二批等。详见说明。 "
            "4.招生人数：仅能填写数字 "
            "5.最高分、最低分、平均分：仅能填写数字，保留小数后两位 "
            "6.一级层次：限定\"本科、专科（高职）\" "
            "7.最低分位次：仅能填写数字 "
            "8.数据来源：必须限定——官方考试院、大红本数据、学校官网、销售、抓取、圣达信、优志愿、学业桥 "
            "9.选科要求：不限科目专业组;多门选考;单科、多科均需选考 "
            "10.选科科目必须是科目的简写（物、化、生、历、地、政、技）"
        )
        
        ws.append([remark_text])
        
        # 第2行：招生年份
        ws.append(['招生年份', admission_year])
        
        # 第3行：表头
        headers = [
            '学校名称', '省份', '招生专业', '专业方向（选填）', '专业备注（选填）',
            '一级层次', '招生科类', '招生批次', '招生类型（选填）', '最高分',
            '最低分', '平均分', '最低分位次（选填）', '招生人数（选填）',
            '数据来源', '专业组代码', '首选科目', '选科要求', '次选科目',
            '专业代码', '招生代码', '最低分数区间低', '最低分数区间高',
            '最低分数区间位次低', '最低分数区间位次高', '录取人数（选填）'
        ]
        ws.append(headers)
        
        # 数据行
        for row_data in converted_data:
            row_values = [row_data.get(header, '') for header in headers]
            ws.append(row_values)
        
        # 合并第一行
        ws.merge_cells('A1:Y1')
        ws['A1'].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        ws.row_dimensions[1].height = 100
        
        # 设置列宽
        for col_idx, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
        
        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in ws[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 保存到字节流
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    except Exception as e:
        logger.error(f"转换导出失败: {e}")
        raise
